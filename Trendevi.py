#Doğrulama Kodu
import requests
from bs4 import BeautifulSoup
url = "https://docs.google.com/spreadsheets/d/1AP9EFAOthh5gsHjBCDHoUMhpef4MSxYg6wBN0ndTcnA/edit#gid=0"
response = requests.get(url)
html_content = response.content
soup = BeautifulSoup(html_content, "html.parser")
first_cell = soup.find("td", {"class": "s2"}).text.strip()
if first_cell != "Aktif":
    exit()
first_cell = soup.find("td", {"class": "s1"}).text.strip()
print(first_cell)


import requests
import pandas as pd
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
from openpyxl.worksheet.table import Table, TableStyleInfo
import shutil
import datetime
import zipfile
from colorama import init, Fore, Style
from selenium.webdriver.chrome.service import Service
import tkinter as tk
from tkinter import simpledialog
import chromedriver_autoinstaller

pd.options.mode.chained_assignment = None





init(autoreset=True)

print(" ")
print(Fore.BLUE + "https://www.modaymis.com/admin/exportorder/edit/20")


print(" ")
print("Oturum Açma Başarılı Oldu")
print(" /﹋\ ")
print("(҂`_´)")
print("<,︻╦╤─ ҉ - -")
print("/﹋\\")
print("(Kod Bekçisi)")
print("Mustafa ARI")
print(" ")






# Verilen URL'den Excel dosyasını indir
url = "https://www.modaymis.com/FaprikaOrderXls/TQEH56/1/"
response = requests.get(url)
excel_content = response.content

# Excel dosyasını bir veri çerçevesine yükle
df = pd.read_excel(excel_content)

# İlk sütundaki "Id" değerlerini benzersiz olarak al
unique_ids = df["Id"].unique()

chromedriver_autoinstaller.install()
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('--log-level=1') 
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])  
driver = webdriver.Chrome(options=chrome_options)

login_url = "https://www.modaymis.com/kullanici-giris/?ReturnUrl=%2Fadmin"
driver.get(login_url)

email_input = driver.find_element("id", "EmailOrPhone")
email_input.send_keys("mustafa@modaymis.com")

password_input = driver.find_element("id", "Password")
password_input.send_keys("123456")
password_input.send_keys(Keys.RETURN)

# Selenium ile giriş yapma işlemi tamamlandı, şimdi "Id" değerlerini işleyebiliriz
for id_value in unique_ids:
    order_url = f"https://www.modaymis.com/admin/order/edit/{id_value}"
    driver.get(order_url)
    
    # Sayfanın yüklenmesini beklemek için WebDriverWait kullanımı
    wait = WebDriverWait(driver, 10)
    note_element = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "note-warning")))
    
    # JavaScript kodunu çalıştırma
    javascript_code = """
        // // İlgili HTML içeriğini seçiyoruz
var noteElement = document.querySelector('.note-warning');
var inputElement = document.querySelector('#TrackingNumber');

// Eğer gerekli elementler bulunduysa ve içerik istediğiniz formatta ise işleme devam ediyoruz
if (noteElement && inputElement) {
    var innerHTML = noteElement.innerHTML;

    // "Kampanya kodu" bilgisini çıkarmak için bir düzenli ifade (regex) kullanabiliriz
    var regex = /Kampanya kodu:\s*(\d+)/;
    var match = innerHTML.match(regex);

    if (match) {
        var kampanyaKodu = match[1];

        // Kampanya kodunu input alanına yapıştırıyoruz
        inputElement.value = kampanyaKodu;
        console.log('Kampanya Kodu:', kampanyaKodu, 'Input alana yapıştırıldı:', inputElement.value);
    } else {
        console.log('Kampanya kodu bulunamadı.');
    }
} else {
    console.log('İlgili element(ler) bulunamadı.');
}

// Belirtilen butonu seçiyoruz
var saveButton = document.querySelector('#btnSaveTrackingNumber');

// Eğer buton bulunduysa otomatik olarak tıklıyoruz
if (saveButton) {
    saveButton.click();
    console.log('Kaydet butonuna otomatik olarak tıklandı.');
} else {
    console.log('Kaydet butonu bulunamadı.');
}

// Belirtilen ikinci butonu seçiyoruz
var confirmButton = document.querySelector('#btnSaveTrackingNumber-action-confirmation-submit-button');

// Eğer ikinci buton bulunduysa otomatik olarak tıklıyoruz
if (confirmButton) {
    confirmButton.click();
    console.log('Onay butonuna otomatik olarak tıklandı.');
} else {
    console.log('Onay butonu bulunamadı.');
}
        
    """
    driver.execute_script(javascript_code)

# Selenium işlemlerinin ardından tarayıcıyı kapatma
driver.quit()


















# GET isteği gönderilecek link
url = "https://www.modaymis.com/FaprikaOrderXls/TQEH56/1/"

# GET isteği gönderme
response = requests.get(url)

# Gelen içeriği bir Excel dosyası olarak kaydetme
with open("veri.xlsx", "wb") as f:
    f.write(response.content)

# Excel dosyasını okuma
df = pd.read_excel("veri.xlsx")

# "Adet" sütunundaki metin verileri sayılara dönüştürme
def convert_to_numeric(value):
    try:
        return float(value.replace(",", "."))
    except ValueError:
        return value

df["Adet"] = df["Adet"].apply(convert_to_numeric)

# İşlemi gerçekleştiren fonksiyon
def duplicate_rows(row):
    count = int(row["Adet"])
    return pd.concat([row] * count, axis=1).T

# Tüm satırları işleme tabi tutma
new_rows = df.apply(duplicate_rows, axis=1)

# Yeni veri çerçevesini oluşturma
new_df = pd.concat(new_rows.tolist(), ignore_index=True)

# Sadece belirtilen sütunları seçme
selected_columns = ["Id", "Barkod", "UrunAdi", "Varyant"]
new_df = new_df[selected_columns]

# Veriyi yeni bir Excel dosyasına yazma
new_df.to_excel("sonuc.xlsx", index=False)









url = "https://haydigiy.online/Products/rafkodlari.php"
response = requests.get(url)
soup = BeautifulSoup(response.text, "html.parser")
table = soup.find("table")
data = []
for row in table.find_all("tr"):
    row_data = []
    for cell in row.find_all(["th", "td"]):
        row_data.append(cell.get_text(strip=True))
    data.append(row_data)
df = pd.DataFrame(data[1:], columns=data[0])
df.to_excel("Raf Kodu.xlsx", index=False)






# "sonuc.xlsx" ve "Raf Kodu.xlsx" dosyalarını okuma
sonuc_df = pd.read_excel("sonuc.xlsx")
google_sheet_df = pd.read_excel("Raf Kodu.xlsx")

# "sonuc.xlsx" dosyasına yeni bir sütun ekleyerek başlangıçta "Raf Kodu Yok" değerleri ile doldurma
sonuc_df["GoogleSheetVerisi"] = "Raf Kodu Yok"

# Her bir "Barkod" değeri için işlem yapma
for index, row in sonuc_df.iterrows():
    barkod = row["Barkod"]
    
    # "Raf Kodu.xlsx" dosyasında ilgili "Barkod"u arama
    matching_row = google_sheet_df[google_sheet_df.iloc[:, 0] == barkod]
    
    # Eşleşen "Barkod" varsa ve karşılık gelen hücre boş değilse, değeri "GoogleSheetVerisi" sütununa yazma
    if not matching_row.empty and not pd.isnull(matching_row.iloc[0, 2]):
        sonuc_df.at[index, "GoogleSheetVerisi"] = matching_row.iloc[0, 2]

# "sonuc.xlsx" dosyasını güncelleme
sonuc_df.to_excel("sonuc.xlsx", index=False)







# "sonuc.xlsx" dosyasını güncelleme
sonuc_df["GoogleSheetVerisi Kopya"] = sonuc_df["GoogleSheetVerisi"]  # "GoogleSheetVerisi" sütununu kopyala
sonuc_df["GoogleSheetVerisi Kopya"] = sonuc_df["GoogleSheetVerisi Kopya"].str.split("-", n=1).str[0]  # "-" den sonrasını temizle
sonuc_df["GoogleSheetVerisi Kopya"] = pd.to_numeric(sonuc_df["GoogleSheetVerisi Kopya"], errors="coerce")  # Sayıya dönüştür
sonuc_df = sonuc_df.sort_values(by="GoogleSheetVerisi Kopya")  # "GoogleSheetVerisi Kopya" sütununa göre sırala

# "GoogleSheetVerisi Kopya" sütununu sil
sonuc_df.drop("GoogleSheetVerisi Kopya", axis=1, inplace=True)

sonuc_df.to_excel("sonuc.xlsx", index=False)









import os

# Dosyaların adları
excel_files_to_delete = ["veri.xlsx", "Raf Kodu.xlsx"]

# Dosyaları sil
for file_name in excel_files_to_delete:
    if os.path.exists(file_name):
        os.remove(file_name)
        
    else:
        print(f"{file_name} dosyası bulunamadı.")










# "sonuc.xlsx" dosyasını güncelleme
sonuc_df["UrunAdi Kopya"] = sonuc_df["UrunAdi"]  # "UrunAdi" sütununu kopyala
sonuc_df["UrunAdi Kopya"] = sonuc_df["UrunAdi Kopya"].str.split("-", n=1).str[1]  # "-" den öncesini temizle

# "UrunAdi Kopya" sütununu "sonuc.xlsx" dosyasına ekleyerek güncelleme
with pd.ExcelWriter("sonuc.xlsx") as writer:
    sonuc_df.to_excel(writer, index=False)










# "UrunAdi" sütununu en sağına yapıştırma
sonuc_df["UrunAdiKopya2"] = sonuc_df["UrunAdi"]

# "UrunAdiKopya2" sütununda " - " dan sonrasını ve son boşluktan öncesini silme
sonuc_df["UrunAdiKopya2"] = sonuc_df["UrunAdiKopya2"].apply(lambda x: x.split(" - ")[0].rsplit(" ", 1)[-1].strip() if " - " in x else x)

# "UrunAdiKopya2" sütununu en sağa taşıma
column_order = list(sonuc_df.columns)
column_order.remove("UrunAdiKopya2")
column_order.append("UrunAdiKopya2")
sonuc_df = sonuc_df[column_order]

# "sonuc.xlsx" dosyasını güncelleme
sonuc_df.to_excel("sonuc.xlsx", index=False)







# "UrunAdi" sütununu tablonun sonuna bir kez daha kopyalama ve düzenleme
sonuc_df["UrunAdiKopya3"] = sonuc_df["UrunAdi"].apply(lambda x: x.split(" - ", 1)[0].strip() if " - " in x else x)

# "UrunAdiKopya3" sütununu en sağa taşıma
column_order = list(sonuc_df.columns)
column_order.remove("UrunAdiKopya3")
column_order.append("UrunAdiKopya3")
sonuc_df = sonuc_df[column_order]

# "sonuc.xlsx" dosyasını güncelleme
sonuc_df.to_excel("sonuc.xlsx", index=False)










# Verileri birleştirip yeni sütun oluşturma
sonuc_df["BirlesikVeri"] = sonuc_df["UrunAdi Kopya"] + " - " + sonuc_df["UrunAdiKopya2"] + " - " + sonuc_df["Varyant"]

# "BirlesikVeri" sütununu en sağa taşıma
column_order = list(sonuc_df.columns)
column_order.remove("BirlesikVeri")
column_order.append("BirlesikVeri")
sonuc_df = sonuc_df[column_order]

# "sonuc.xlsx" dosyasını güncelleme
sonuc_df.to_excel("sonuc.xlsx", index=False)


# "BirlesikVeri" sütunundaki "Beden:" ibaresini çıkarma
sonuc_df["BirlesikVeri"] = sonuc_df["BirlesikVeri"].str.replace("Beden:", "")

# "sonuc.xlsx" dosyasını güncelleme
sonuc_df.to_excel("sonuc.xlsx", index=False)






# Belirtilen sütunları silme
columns_to_drop = ["UrunAdi", "Varyant", "UrunAdi Kopya", "UrunAdiKopya2"]
sonuc_df.drop(columns_to_drop, axis=1, inplace=True)

# "sonuc.xlsx" dosyasını güncelleme
sonuc_df.to_excel("sonuc.xlsx", index=False)







# "Id" sütununu teke düşürme
unique_ids = sonuc_df["Id"].drop_duplicates()

# Yeni bir Excel sayfası oluşturma ve "Id" değerlerini yazma
with pd.ExcelWriter("sonuc.xlsx", engine="openpyxl", mode="a") as writer:
    unique_ids.to_excel(writer, sheet_name="Unique Ids", index=False)











from openpyxl import load_workbook

# Sonuç dosyasını yükle
file_path = "sonuc.xlsx"
wb = load_workbook(file_path)
sheet = wb["Unique Ids"]

# Başlangıç sütunu ve satırı
start_column = 2
start_row = 2

# Toplam tekrar sayısı ve her tekrardaki numara adedi
repeat_count = 50
numbers_per_repeat = 28

# Verileri ekleme
for _ in range(repeat_count):
    for num in range(1, numbers_per_repeat + 1):
        sheet.cell(row=start_row, column=start_column, value=num)
        start_row += 1

# Değişiklikleri kaydetme
wb.save("sonuc.xlsx")







# Sonuç dosyasını yükle
file_path = "sonuc.xlsx"
wb = load_workbook(file_path)
sheet = wb["Unique Ids"]

# Başlangıç sütunu ve satırı
start_column = 3
start_row = 2

# Toplam tekrar sayısı ve her tekrardaki numara adedi
repeat_count = 50
numbers_per_repeat = 28

# Verileri ekleme
for repeat in range(repeat_count):
    for num in range(1, numbers_per_repeat + 1):
        sheet.cell(row=start_row, column=start_column, value=(repeat % numbers_per_repeat) + 1)
        start_row += 1

# Değişiklikleri kaydetme
wb.save("sonuc.xlsx")











# Sonuç dosyasını yükle
file_path = "sonuc.xlsx"
wb = load_workbook(file_path)
unique_ids_sheet = wb["Unique Ids"]
main_sheet = wb["Sheet1"]

# "Id" sütununun verilerini al
id_column = main_sheet["A"][1:]
unique_ids_column = unique_ids_sheet["A"][1:]

# Karşılık gelen değerleri bulup "Sheet1" sayfasının en sağında yeni bir sütuna ekle
new_column = main_sheet.max_column + 1
main_sheet.cell(row=1, column=new_column, value="Matching Value")

for id_cell in id_column:
    id_value = id_cell.value
    for unique_id_cell in unique_ids_column:
        if unique_id_cell.value == id_value:
            matching_value = unique_id_cell.offset(column=1).value
            main_sheet.cell(row=id_cell.row, column=new_column, value=matching_value)
            break

# Değişiklikleri kaydetme
wb.save("sonuc.xlsx")












# Sonuç dosyasını yükle
file_path = "sonuc.xlsx"
wb = load_workbook(file_path)
unique_ids_sheet = wb["Unique Ids"]
main_sheet = wb["Sheet1"]

# "Id" sütununun verilerini al
id_column = main_sheet["A"][1:]
unique_ids_column = unique_ids_sheet["A"][1:]

# Karşılık gelen değerleri bulup "Sheet1" sayfasının en sağında yeni bir sütuna ekle
new_column = main_sheet.max_column + 1
main_sheet.cell(row=1, column=new_column, value="Matching Value (3rd Column)")

for id_cell in id_column:
    id_value = id_cell.value
    for unique_id_cell in unique_ids_column:
        if unique_id_cell.value == id_value:
            matching_value = unique_id_cell.offset(column=2).value
            main_sheet.cell(row=id_cell.row, column=new_column, value=matching_value)
            break

# Değişiklikleri kaydetme
wb.save("sonuc.xlsx")











# Sonuç dosyasını yükle
file_path = "sonuc.xlsx"
wb = load_workbook(file_path)
main_sheet = wb["Sheet1"]

# Sütun başlıklarını değiştir
new_column_titles = {
    "Id": "SiparişNO",
    "BirlesikVeri": "ÜRÜN",
    "GoogleSheetVerisi": "RAF KODU",
    "UrunAdiKopya3": "ÜRÜN ADI",
    "Matching Value": "KUTU",
    "Matching Value (3rd Column)": "ÇN"
}

for col_idx, col_name in enumerate(main_sheet.iter_cols(min_col=1, max_col=main_sheet.max_column), start=1):
    old_title = col_name[0].value
    new_title = new_column_titles.get(old_title, old_title)
    col_name[0].value = new_title

# Değişiklikleri kaydetme
wb.save("sonuc.xlsx")









from openpyxl.utils.dataframe import dataframe_to_rows

# Sonuç dosyasını yükle
file_path = "sonuc.xlsx"
wb = load_workbook(file_path)
main_sheet = wb["Sheet1"]

# Sütunların yeni sıralaması
new_column_order = [
    "RAF KODU",
    "ÜRÜN",
    "Barkod",
    "KUTU",
    "ÜRÜN ADI",
    "ÇN",
    "SiparişNO"
]

# Yeni bir DataFrame oluştur
data = main_sheet.iter_rows(min_row=2, values_only=True)
df = pd.DataFrame(data, columns=[cell.value for cell in main_sheet[1]])

# Sütunları yeni sıralamaya göre düzenle
df = df[new_column_order]

# Mevcut başlıkları güncelle
for idx, column_name in enumerate(new_column_order, start=1):
    main_sheet.cell(row=1, column=idx, value=column_name)

# DataFrame verilerini sayfaya yaz
for r_idx, row in enumerate(df.values, 2):
    for c_idx, value in enumerate(row, 1):
        main_sheet.cell(row=r_idx, column=c_idx, value=value)

# Değişiklikleri kaydet
wb.save("sonuc.xlsx")









from openpyxl.styles import Alignment, Font

# Sonuç dosyasını yükle
file_path = "sonuc.xlsx"
wb = load_workbook(file_path)
main_sheet = wb["Sheet1"]

# Hücreleri ortala ve ortaya hizala
for row in main_sheet.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")


# Değişiklikleri kaydet
wb.save("sonuc.xlsx")








from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side

# Sonuç dosyasını yükle
file_path = "sonuc.xlsx"
wb = load_workbook(file_path)
main_sheet = wb["Sheet1"]

# Kenarlık stili oluştur
border_style = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# Hücreleri kalın yap, yazı fontunu 14 yap ve kenarlık ekle
for row in main_sheet.iter_rows():
    for cell in row:
        cell.font = Font(bold=True, size=14)
        cell.border = border_style

# Değişiklikleri kaydet
wb.save("sonuc.xlsx")




from openpyxl.utils import get_column_letter

# Sonuç dosyasını yükle
file_path = "sonuc.xlsx"
wb = load_workbook(file_path)
main_sheet = wb["Sheet1"]

# "RAF KODU" sütununu 45 piksel yap
main_sheet.column_dimensions["C"].width = 45

# Tüm hücreleri en uygun sütun genişliği olarak ayarla
for column in main_sheet.columns:
    max_length = 0
    column = list(column)
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    main_sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

# Değişiklikleri kaydet
wb.save("sonuc.xlsx")





# Sonuç dosyasını yükle
file_path = "sonuc.xlsx"
wb = load_workbook(file_path)
main_sheet = wb["Sheet1"]

# İlk sütunu (A sütunu) 45 piksel genişliğinde yap
main_sheet.column_dimensions["A"].width = 45
main_sheet.column_dimensions["C"].width = 14
main_sheet.column_dimensions["G"].width = 14
main_sheet.column_dimensions["D"].width = 9
main_sheet.column_dimensions["F"].width = 5

# Değişiklikleri kaydet
wb.save("sonuc.xlsx")










from copy import copy

# Sonuç dosyasını yükle
file_path = "sonuc.xlsx"
wb = load_workbook(file_path)
main_sheet = wb["Sheet1"]

# Tüm hücrelere "Metni Kaydır" formatını uygula
for row in main_sheet.iter_rows():
    for cell in row:
        new_alignment = copy(cell.alignment)
        new_alignment.wrap_text = True
        cell.alignment = new_alignment

# Değişiklikleri kaydet
wb.save("sonuc.xlsx")









from openpyxl.worksheet.table import Table, TableStyleInfo

# Sonuç dosyasını yükle
file_path = "sonuc.xlsx"
wb = load_workbook(file_path)
main_sheet = wb["Sheet1"]

# Tabloyu oluşturma
table = Table(displayName="MyTable", ref=main_sheet.dimensions)

# Tablo stili oluşturma (gri-beyaz)
style = TableStyleInfo(
    name="TableStyleLight1", showFirstColumn=False,
    showLastColumn=False, showRowStripes=True, showColumnStripes=True
)

# Tabloya stil atama
table.tableStyleInfo = style

# Tabloyu sayfaya ekleme
main_sheet.add_table(table)

# Değişiklikleri kaydetme
wb.save("sonuc.xlsx")














def create_bat_files(data, batch_size=28):
    batch_count = 1
    batch_data = []
    remaining_data = data

    while len(remaining_data) > 0:
        current_batch = remaining_data[:batch_size]
        batch_data.extend(current_batch)

        with open(f"BAT{batch_count}.bat", "w") as file:
            for index, value in enumerate(current_batch):
                link = f"https://www.modaymis.com/admin/order/edit/{value}"
                file.write(f'start "" {link}\n')
                if index == 0:
                    file.write('timeout -t 2\n')  # Add the timeout line only after the first link
        batch_data = []
        remaining_data = remaining_data[batch_size:]
        batch_count += 1


# Sonuç dosyasını yükle
file_path = "sonuc.xlsx"
wb = load_workbook(file_path)
unique_ids_sheet = wb["Unique Ids"]

# "Id" sütunundaki verileri al
id_column = unique_ids_sheet["A"][1:]

# Verileri bir listeye dönüştür
id_values = [cell.value for cell in id_column if cell.value is not None]

# .bat dosyalarını oluştur
create_bat_files(id_values)









import os

# Eski ve yeni dosya adları
old_file_name = "sonuc.xlsx"
new_file_name = "Trendevi.xlsx"

# Dosya adını değiştir
if os.path.exists(old_file_name):
    os.rename(old_file_name, new_file_name)
    
else:
    print(f"{old_file_name} dosyası bulunamadı.")











import os
import shutil

def main():
    # Mevcut dizini al
    current_directory = os.getcwd()

    # Yeni klasör adını belirle
    new_folder_name = "Trendevi"

    # Yeni klasörün tam yolu
    new_folder_path = os.path.join(current_directory, new_folder_name)

    # Klasörü oluştur
    os.makedirs(new_folder_path, exist_ok=True)

    # .bat uzantılı dosyaları ve "Trendevi.xlsx" dosyasını taşı
    move_files(current_directory, new_folder_path)

def move_files(source_folder, destination_folder):
    # Kaynak klasördeki tüm dosyaları al
    files = os.listdir(source_folder)

    for file in files:
        if file.endswith(".bat") or file == "Trendevi.xlsx":
            # Dosyanın tam yolu
            source_file_path = os.path.join(source_folder, file)

            # Hedef klasördeki dosyanın tam yolu
            destination_file_path = os.path.join(destination_folder, file)

            # Dosyayı taşı
            shutil.move(source_file_path, destination_file_path)

if __name__ == "__main__":
    main()




    
import datetime
import zipfile
folders = ["Trendevi"]

# Bugünkü tarihi al
current_date = datetime.datetime.now().strftime("%Y-%m-%d")

# Oluşturulacak zip dosyasının adı
zip_filename = f"{current_date} Trendevi.zip"

# Klasörleri kontrol et ve gerektiğinde sil veya zip'e ekle
with zipfile.ZipFile(zip_filename, 'w') as zipf:
    for folder in folders:
        folder_path = os.path.join(".", folder)
        folder_contents = os.listdir(folder_path)
        bat_files = [file for file in folder_contents if file.endswith(".bat")]

        if bat_files:
            for root, _, files in os.walk(folder_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    zipf.write(file_path, os.path.relpath(file_path, "."))
            
            
        else:
            for root, dirs, files in os.walk(folder_path, topdown=False):
                for file in files:
                    file_path = os.path.join(root, file)
                    os.remove(file_path)
                for dir in dirs:
                    dir_path = os.path.join(root, dir)
                    os.rmdir(dir_path)
            os.rmdir(folder_path)



# Klasörleri sil
for folder in folders:
    folder_path = os.path.join(".", folder)
    if os.path.exists(folder_path):
        shutil.rmtree(folder_path)